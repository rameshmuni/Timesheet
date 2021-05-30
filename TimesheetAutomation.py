
from openpyxl import load_workbook
from datetime import datetime
import time
import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Color, PatternFill, Font, Border
from datetime import date, timedelta
from openpyxl.styles import Border, Side

def ReadColrange(rm,cm,ws):
    for i in range(1,rm+1):
        for j in range(1,cm+1):
            if i==2 and j>5:
                val=ws.cell(row=i,column=j).value
                if val!=None:
                    d=str(val).split(" ")[0]
                    if d!='Day':

                        if int(d.split("-")[1])==int(now):

                            colrange.append(j)
    return colrange

def FillTemplate(ws,rm,wbr,colrange,sheet):
    for i in range(4,rm+1):
        ResourceName=ws.cell(row=i,column=4).value
        ResourceID=ws.cell(row=i,column=1).value
        PONumber=ws.cell(row=i,column=2).value
        Role=ws.cell(row=i,column=3).value
        r=12
        islastday=False
        daycount=1
        hourscount=0
        if ResourceID!=None:
            wsr=wbr.create_sheet(ResourceName)
        for j in range(int(colrange[0]),int(colrange[1])):
            val=ws.cell(row=i,column=j).value
            try:
                cmt = ws.cell(row=i,column=j).comment.text
            except:
                cmt=""       
            if ResourceID!=None:
                
                try:
                    if int(val)>=0:
                        islastday=True
                        wsr.cell(row=43,column=7).value=int(val)*8
                        wsr.cell(row=8,column=7).value=val
                        wsr.cell(row=8,column=9).value=int(val)*8
                except:
                    pass
                if not islastday:
                    datetoenter=str(daycount)+"/"+m+"/"+y
                    weeknumber=datetime.date(int(y), int(m), daycount).isocalendar()[1]
                    wsr.cell(row=r,column=2).value="W"+str(weeknumber)
                    wsr.merge_cells('C'+str(r)+':E'+str(r))
                    cell = wsr.cell(row=r,column=3)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    wsr.cell(row=r,column=3).value=activityMap[val]
                    if activityMap[val]=='WO':
                        Fill6 = PatternFill(start_color='00808080', end_color='00808080', fill_type='solid')
                        fillingcolor(Fill6, r, 3, 6, wsr, wbr)

                    wsr.cell(row=r,column=6).value=datetoenter
                    wsr.cell(row=r,column=7).value=Timings[val][2]
                    wsr.cell(row=r,column=8).value=Timings[val][0]
                    wsr.cell(row=r,column=9).value=Timings[val][1]
                    wsr.cell(row=r,column=10).value=Timings[val][1]
                    wsr.cell(row=r,column=11).value='Bangalore'
                    wsr.merge_cells('L'+str(r)+':N'+str(r))
                    cell = wsr.cell(row=r,column=12)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    wsr.cell(row=r,column=12).value=cmt
                    r+=1
                    daycount+=1
                    hourscount+=int(Timings[val][2])
        GenerateTemplate(wsr,wbr,sheet,PONumber,ResourceName,ResourceID,Role)
        wbr.save('output.xlsx')
        
def GenerateTemplate(wsr,wbr,sheet,PONumber,ResourceName,ResourceID,Role):
    print(PONumber,ResourceName,ResourceID,Role)
    wsr.merge_cells('B2:N2')
    cell = wsr.cell(row=2,column=2)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value='Airbus India- ITSM Project Timesheet'
    wsr.cell(row=3,column=2).value='Project'
    
    wsr.merge_cells('C3:F3')
    cell = wsr.cell(row=3,column=3)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value='ITSM'
    
    wsr.cell(row=4,column=2).value='Supplier'
    
    wsr.merge_cells('C4:F4')
    cell = wsr.cell(row=4,column=3)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value=sheet
    wbr.save('output.xlsx')
    wsr.cell(row=4,column=7).value='PO'
    
    wsr.merge_cells('H4:N4')
    cell = wsr.cell(row=4,column=8)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value=PONumber
    
    wsr.cell(row=3,column=7).value='Month'
    
    wsr.merge_cells('H3:N3')
    cell = wsr.cell(row=3,column=8)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value=monthname
    
    wsr.merge_cells('B6:C6')
    cell = wsr.cell(row=6,column=2)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value='Name'
    
    wsr.merge_cells('D6:E6')
    cell = wsr.cell(row=6,column=4)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value=ResourceName

    wsr.cell(row=6, column=4).value = ResourceName
    
    wsr.merge_cells('B7:C7')
    cell = wsr.cell(row=7,column=2)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value='Designation'
    
    wsr.merge_cells('D7:E7')
    cell = wsr.cell(row=7,column=4)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value=Role
    
    wsr.merge_cells('B8:C8')
    cell = wsr.cell(row=8,column=2)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value='AIRBUS LoginID'
    
    wsr.merge_cells('D8:E8')
    cell = wsr.cell(row=8,column=4)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value=ResourceID
    
    wsr.merge_cells('F6:G6')
    cell = wsr.cell(row=6,column=6)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value='Start date'
    
    wsr.merge_cells('H6:N6')
    cell = wsr.cell(row=6,column=8)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value=start_day_of_prev_month
    
    wsr.merge_cells('F7:G7')
    cell = wsr.cell(row=7,column=6)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value='End date'
    
    wsr.merge_cells('H7:N7')
    cell = wsr.cell(row=7,column=8)                  
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value=last_day_of_prev_month
    
    wsr.cell(row=8,column=8).value='Total Hrs'
    wsr.cell(row=43,column=6).value='Total Hrs'
    wsr.cell(row=8, column=6).value = 'Total Days'
    wsr.cell(row=11, column=2).value = 'Week'

    wsr.merge_cells('C11:E11')
    cell = wsr.cell(row=11, column=3)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = 'Activity'

    wsr.cell(row=11, column=6).value = 'Date'
    wsr.cell(row=11, column=7).value = 'Hours'
    wsr.cell(row=11, column=8).value = 'Office In Time'
    wsr.cell(row=11, column=9).value = 'Office Out Time'
    wsr.cell(row=11, column=10).value = 'Final Work End Time'
    wsr.cell(row=11, column=11).value = 'Location'

    wsr.merge_cells('L11:N11')
    cell = wsr.cell(row=11, column=12)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = 'Remarks'

    Fill5 = PatternFill(start_color='00FFCC99', end_color='00FFCC99', fill_type='solid')
    fillingcolor(Fill5, 11,2,14,wsr,wbr)

    colorranges(wsr,wbr)
    wbr.save('output.xlsx')
def numtochar(colstart,colend):
    alphabet=list('abcdefghijklmnopqrstuvwxyz')
    res=[]
    for i in range(colstart,colend):
        res.append(alphabet[i-1])
    return res
def fillingcolor(Fill,rownum,colstart,colend,wsr,wbr):
    res=[]
    res=numtochar(colstart, colend)
    for i in res:
        wsr[i+str(rownum)].fill=Fill
    wbr.save('output.xlsx')

def colorranges(wsr,wbr):
    Fill1 = PatternFill(start_color='009999FF',end_color='009999FF',fill_type='solid')
    Fill2 = PatternFill(start_color='00FFFF99', end_color='00FFFF99', fill_type='solid')
    Fill3 = PatternFill(start_color='0099CCFF', end_color='0099CCFF', fill_type='solid')
    Fill4 = PatternFill(start_color='00FFFF00', end_color='00FFFF00', fill_type='solid')

    fillingcolor(Fill1,2,2,14,wsr,wbr)
    fillingcolor(Fill2, 4, 2, 14, wsr, wbr)
    fillingcolor(Fill2, 3, 2, 14, wsr, wbr)
    fillingcolor(Fill3, 6, 2, 14, wsr, wbr)
    fillingcolor(Fill3, 7, 2, 14, wsr, wbr)
    fillingcolor(Fill3, 8, 2, 15, wsr, wbr)
    fillingcolor(Fill4, 8, 8, 10, wsr, wbr)
    fillingcolor(Fill4, 43, 6, 8, wsr, wbr)

    '''Filling table borders'''
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in wsr['B11:N42']:
        for cell in row:
            cell.border = border
    for row in wsr['B6:N8']:
        for cell in row:
            cell.border = border

    for row in wsr['B2:N4']:
        for cell in row:
            cell.border = border

    wbr.save('output.xlsx')

wb=load_workbook("MasterSheet.xlsx")
worksheets=wb.sheetnames
ws=wb[worksheets[0]]
rm=ws.max_row
cm=ws.max_column
now = str(datetime.datetime.now()).split(" ")[0]
print(now)
now=str(int(now.split("-")[1])-1)
last_month = datetime.datetime.now() - relativedelta(months=1)
monthname = format(last_month, '%B')
last_day_of_prev_month = date.today().replace(day=1) - timedelta(days=1)
start_day_of_prev_month = date.today().replace(day=1) - timedelta(days=last_day_of_prev_month.day)
today = datetime.date.today()
first = today.replace(day=1)
lastMonth = first - datetime.timedelta(days=1)
y=lastMonth.strftime("%Y")
m=lastMonth.strftime("%m")
wb1=load_workbook("Template.xlsx")
worksheets1=wb1.sheetnames
ws1=wb1[worksheets1[0]]
wbr=Workbook()
templatevalues={'header':'B2','Project':'C3','Supplier':'C4'}
activityMap={'S1':'Workedin Airbus','S2':'Worked in Airbus',"HW":"Holiday Work","L":"Leave","CO":"Compoff","WFH":"Worked in airbus",'IH':'Indian Holiday',None:'WO'}
Timings={'S1':('10:30','18:30','8'),'S2':('13:30','22:30','8'),'WFH':('0','0','0'),'CO':('0','0','0'),'IH':('0','0','0'),'L':('0','0','0'),'HW':('0','0','8'),None:('0','0','0')}

for sheet in worksheets:
    ws=wb[sheet]
    rm=ws.max_row
    cm=ws.max_column
    colrange=[]
    colrange=ReadColrange(rm,cm,ws)
    print(colrange)
    FillTemplate(ws,rm,wbr,colrange,sheet)
    time.sleep(15)

                
                    
                    
            
            

